import sys, os
if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QWidget,QFileDialog,QMessageBox,QDesktopWidget,QAbstractItemView
from PyQt5.QtGui import QPalette,QPixmap,QBrush,QIcon
from PyQt5.QtCore import QTimer, Qt
from Utils.ConfigRead import *
import os.path, re,time
import shutil
from openpyxl import load_workbook
from Utils.WriteFile import YamlWrite
from Utils.ParseYaml import ParseYaml
import threading
import win32api
import win32con
from threading import Timer
from testcase.testPaperless import TestPaperless
from Utils.ParseExcel import ParseExcel
from selenium import webdriver


class Application(QWidget):

    def __init__(self):
        super(Application, self).__init__()
        self.timer = QTimer(self)
        # self.schedule = sched.scheduler(time.time, time.sleep)
        # # 修改注册表
        key = win32api.RegOpenKey(win32con.HKEY_CURRENT_USER,
                                  'Software\\Microsoft\\Windows\\Windows Error Reporting', 0, win32con.KEY_ALL_ACCESS)
        if win32api.RegQueryValueEx(key, 'DontShowUI')[0] == 0:
            win32api.RegSetValueEx(key, 'DontShowUI', 0, win32con.REG_DWORD, 1)
            win32api.RegCloseKey(key)
        self.parseyaml = ParseYaml()
        # 模板导出数量计算
        self.filenum = 1
        # 编写用例说明导出数量计算
        self.explainum = 1
        # output定时器执行数计算
        self.OutPut_num = 0
        # case_time定时器执行数计算
        self.CaseTime_Num = 0
        # 导入模板路径
        self.filepath = ''
        # 用例总数集合
        self.CaseNum = []
        # 用例模块数量集合
        self.CaseList = []
        self.initUI()
        self.newfile = ''
        # 如果parameter.py中有IP地址，则键入
        if self.parseyaml.ReadParameter('IP') is not None:
            if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$", self.parseyaml.ReadParameter('IP')):
                self.lineEdit.setText(self.parseyaml.ReadParameter('IP'))

    def initUI(self):
        self.setFixedSize(500, 800)
        self.setWindowFlags(QtCore.Qt.WindowStaysOnTopHint)
        screen = QDesktopWidget().screenGeometry()
        self.setGeometry((screen.width() - 500) / 2, (screen.height() - 800) / 2, 500, 800)
        self.setWindowTitle('自动化测试脚本')
        self.setWindowIcon(QIcon(RESOURSE_PATH + '111.png'))
        palette = QPalette()
        palette.setBrush(QPalette.Background, QBrush(QPixmap(RESOURSE_PATH + '背景色.png')))
        self.setPalette(palette)

        # IP地址位置
        self.label = QtWidgets.QLabel('IP地址', self)
        self.label.setGeometry(QtCore.QRect(40, 15, 83, 30))
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # 浏览器位置
        self.label_2 = QtWidgets.QLabel("浏览器",self)
        self.label_2.setGeometry(QtCore.QRect(40, 60, 83, 30))
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # 测试报告位置
        self.label_3 = QtWidgets.QLabel("测试报告",self)
        self.label_3.setGeometry(QtCore.QRect(40, 105, 83, 30))
        self.label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.label_3.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # 循环位置位置
        self.label_4 = QtWidgets.QLabel("循环次数",self)
        self.label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.label_4.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # 测试用例位置
        self.label_5 = QtWidgets.QLabel("测试用例",self)
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.label_5.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # 用例模块位置
        self.label_6 = QtWidgets.QLabel("用例模块",self)
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # 执行用例位置
        self.label_7 = QtWidgets.QLabel("执行用例",self)
        self.label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.label_7.setStyleSheet("QLabel{font:12pt '宋体';color:white}")

        # IP输入框
        self.lineEdit = QtWidgets.QLineEdit(self)
        self.lineEdit.setGeometry(QtCore.QRect(150, 15, 240, 30))
        self.lineEdit.setStyleSheet("QLineEdit{font:12pt '宋体';border-radius:3px;background-color:rgba(113,113,113,0.5);color:white}")

        # 浏览器选择框
        self.comboBox = QtWidgets.QComboBox(self)
        self.comboBox.setGeometry(QtCore.QRect(150, 60, 240, 30))
        self.comboBox.setStyleSheet(
                                    "QComboBox{font:12pt '宋体'}"
                                    "QComboBox{background-color:rgba(113,113,113,0.5)}"
                                    "QComboBox{color:white}")
        # 加入浏览器
        self.comboBox.addItems([self.getBrowserType()[0], self.getBrowserType()[1]])


        # 生成报告按钮
        self.radioButton = QtWidgets.QRadioButton("生成",self)
        self.radioButton.setGeometry(QtCore.QRect(150, 105, 55, 30))
        self.radioButton.setStyleSheet("QRadioButton{font:12pt '宋体';color:white}")
        # 生成报告点击事件
        self.radioButton.clicked.connect(lambda: self.btnstate(self.radioButton))

        # 不生成报告按钮
        self.radioButton_2 = QtWidgets.QRadioButton("不生成",self)
        self.radioButton_2.setGeometry(QtCore.QRect(240, 105, 70, 30))
        self.radioButton_2.setStyleSheet("QRadioButton{font:12pt '宋体';color:white}")
        # 不生成报告点击事件
        self.radioButton_2.clicked.connect(lambda: self.btnstate(self.radioButton_2))
        # 默认选中不生成测试报告
        self.radioButton_2.setChecked(True)

        # 编写用例说明按钮
        self.toolButton_5 = QtWidgets.QToolButton(self)
        self.toolButton_5.setText("编写说明")
        self.toolButton_5.setStyleSheet(
            "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
        """点击事件"""
        self.toolButton_5.clicked.connect(self.explainclick)

        # 生成报告提示语
        self.label_8 = QtWidgets.QLabel("运行中请勿打开测试报告",self)
        self.label_8.setGeometry(QtCore.QRect(320, 105, 135, 30))
        self.label_8.setStyleSheet("QLabel{font:9pt '宋体';color:red}")
        # 默认选中不生成测试报告,隐藏提示
        self.label_8.setHidden(True)

        # 生成报告路径显示
        self.label_9 = QtWidgets.QLabel('',self)
        self.label_9.setGeometry(QtCore.QRect(150, 135, 290, 20))
        self.label_9.setStyleSheet("QLabel{font:9pt '宋体';color:#97FF48}")
        # 默认选中不生成测试报告，隐藏报告路径显示
        self.label_9.setHidden(True)

        # 执行用例循环测试选择框
        self.spinBox = QtWidgets.QSpinBox(self)
        self.spinBox.setStyleSheet(
                    "QSpinBox{font:12pt '宋体'}"
                    "QSpinBox{border-radius:3px}"
                    "QSpinBox{background-color:rgba(99,99,99,1)}"
                    "QSpinBox{color:white}")
        self.spinBox.setRange(1, 5)  # 设置下界和上界
        # 不能输入值
        self.spinBox.setFocusPolicy(True)


        # 导入测试用例按钮
        # self.toolButton = QtWidgets.QToolButton("导入",None)
        self.toolButton  =  QtWidgets.QToolButton(self)
        self.toolButton.setText('导入')
        self.toolButton.setStyleSheet(
            "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
        """点击事件"""
        self.toolButton.clicked.connect(self.importclick)

        # 导出用例模板按钮
        self.toolButton_2 = QtWidgets.QToolButton(self)
        self.toolButton_2.setText("导出模板")
        self.toolButton_2.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
        """点击事件"""
        self.toolButton_2.clicked.connect(self.exportclick)

        # 导入用例文件显示
        self.label_10 = QtWidgets.QLabel("请选择excel.xlsx格式表格导入", self)
        self.label_10.setStyleSheet("QLabel{font:9pt '宋体';color:#97FF48}")

        # 用例模块选择框
        self.comboBox_2 = QtWidgets.QComboBox(self)
        self.comboBox_2.setStyleSheet(
                "QComboBox{font:12pt '宋体'}"
                "QComboBox{background-color:rgba(113,113,113,0.5)}"
                "QComboBox{color:white}")
        # 加入模块
        self.comboBox_2.addItem('全部')
        self.comboBox_2.currentIndexChanged.connect(self.ComboxValue)

        # 执行用例启动按钮
        self.toolButton_3 = QtWidgets.QToolButton(self)
        self.toolButton_3.setText('启动')
        self.toolButton_3.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
        self.toolButton_3.clicked.connect(self.RunTest)

        # 执行用例暂停按钮
        self.toolButton_4 = QtWidgets.QToolButton(self)
        self.toolButton_4.setText('暂停')
        self.toolButton_4.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
        self.toolButton_4.setEnabled(False)
        self.toolButton_4.clicked.connect(self.Suspend)

        self.label_11 = QtWidgets.QLabel("总用例数量:",self)
        self.label_11.setStyleSheet("QLabel{font:10pt '宋体';color: #FFF666}")
        self.line = QtWidgets.QFrame(self)
        self.line.setStyleSheet("QFrame{background-color:#FFFFFF}")
        # 总用例数量显示
        self.label_15 = QtWidgets.QLabel("0",self)
        self.label_15.setAlignment(QtCore.Qt.AlignCenter)
        self.label_15.setStyleSheet("QLabel{font:10pt '宋体';color: #FFF666}")

        self.label_12 = QtWidgets.QLabel("已执行数量:",self)
        self.label_12.setStyleSheet("QLabel{font:10pt '宋体';color: #FFF666}")
        self.line_2 = QtWidgets.QFrame(self)
        self.line_2.setStyleSheet("QFrame{background-color:#FFFFFF}")
        # 已执行用例数量显示
        self.label_17 = QtWidgets.QLabel("0",self)
        self.label_17.setAlignment(QtCore.Qt.AlignCenter)
        self.label_17.setStyleSheet("QLabel{font:10pt '宋体';color: #FFF666}")

        self.label_13 = QtWidgets.QLabel("运行时间:",self)
        self.label_13.setStyleSheet("QLabel{font:10pt '宋体';color: #FFF666}")
        self.line_3 = QtWidgets.QFrame(self)
        self.line_3.setStyleSheet("QFrame{background-color:#FFFFFF}")
        # 运行时间显示
        self.label_18 = QtWidgets.QLabel("0:00:00",self)
        self.label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.label_18.setStyleSheet("QLabel{font:10pt '宋体';color: #FFF666}")

        # 版本号显示
        self.label_19 = QtWidgets.QLabel('版本号: V2.0', self)
        self.label_19.setAlignment(QtCore.Qt.AlignCenter)
        self.label_19.setStyleSheet("QLabel{font:10pt '宋体';color:white}")

        # 查看日志面板
        self.listWidget = QtWidgets.QListWidget(self)
        item = QtWidgets.QListWidgetItem()
        self.listWidget.setStyleSheet("QListWidget{font:10pt '宋体';border-radius:3px;background-color:#282828;color:white}")
        self.listWidget.addItem(item)
        self.listWidget.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.listWidget.itemClicked.connect(self.myListWidgetValue)
        # self.listWidget.grabKeyboard()


        # 不生成测试报告初始位置
        self.initialposition()

    '''不生成测试报告位置'''
    def initialposition(self):
        self.label_4.setGeometry(QtCore.QRect(40, 150, 83, 30))
        self.label_5.setGeometry(QtCore.QRect(40, 195, 83, 30))
        self.label_6.setGeometry(QtCore.QRect(40, 250, 83, 30))
        self.label_7.setGeometry(QtCore.QRect(40, 295, 83, 30))
        self.spinBox.setGeometry(QtCore.QRect(150, 150, 55, 30))
        self.toolButton.setGeometry(QtCore.QRect(150, 195, 70, 30))
        self.toolButton_2.setGeometry(QtCore.QRect(250, 195, 81, 30))
        self.toolButton_5.setGeometry(QtCore.QRect(360, 195, 81, 30))
        self.label_10.setGeometry(QtCore.QRect(150, 227, 280, 20))
        self.comboBox_2.setGeometry(QtCore.QRect(150, 250, 230, 30))
        self.toolButton_3.setGeometry(QtCore.QRect(150, 295, 70, 30))
        self.toolButton_4.setGeometry(QtCore.QRect(250, 295, 70, 30))
        self.label_11.setGeometry(QtCore.QRect(50, 350, 70, 20))
        self.line.setGeometry(QtCore.QRect(125, 367, 105, 1))
        self.label_15.setGeometry(QtCore.QRect(125, 348, 105, 19))
        self.label_12.setGeometry(QtCore.QRect(50, 375, 70, 20))
        self.line_2.setGeometry(QtCore.QRect(125, 392, 105, 1))
        self.label_17.setGeometry(QtCore.QRect(125, 372, 105, 19))
        self.label_13.setGeometry(QtCore.QRect(50, 400, 70, 20))
        self.line_3.setGeometry(QtCore.QRect(125, 417, 105, 1))
        self.label_18.setGeometry(QtCore.QRect(125, 397, 105, 19))
        self.label_19.setGeometry(QtCore.QRect(390, 410, 105, 19))
        self.listWidget.setGeometry(QtCore.QRect(2, 440, 496, 358))

    '''生成测试报告位置'''
    def changelocation(self):
        self.label_4.setGeometry(QtCore.QRect(40, 160, 83, 30))
        self.label_5.setGeometry(QtCore.QRect(40, 205, 83, 30))
        self.label_6.setGeometry(QtCore.QRect(40, 260, 83, 30))
        self.label_7.setGeometry(QtCore.QRect(40, 305, 83, 30))
        self.spinBox.setGeometry(QtCore.QRect(150, 160, 55, 30))
        self.toolButton.setGeometry(QtCore.QRect(150, 205, 70, 30))
        self.toolButton_2.setGeometry(QtCore.QRect(250, 205, 81, 30))
        self.toolButton_5.setGeometry(QtCore.QRect(360, 205, 81, 30))
        self.label_10.setGeometry(QtCore.QRect(150, 237, 280, 20))
        self.comboBox_2.setGeometry(QtCore.QRect(150, 260, 230, 30))
        self.toolButton_3.setGeometry(QtCore.QRect(150, 305, 70, 30))
        self.toolButton_4.setGeometry(QtCore.QRect(250, 305, 70, 30))
        self.label_11.setGeometry(QtCore.QRect(50, 360, 70, 20))
        self.line.setGeometry(QtCore.QRect(125, 377, 105, 1))
        self.label_15.setGeometry(QtCore.QRect(125, 358, 105, 19))
        self.label_12.setGeometry(QtCore.QRect(50, 385, 70, 20))
        self.line_2.setGeometry(QtCore.QRect(125, 402, 105, 1))
        self.label_17.setGeometry(QtCore.QRect(125, 382, 105, 19))
        self.label_13.setGeometry(QtCore.QRect(50, 410, 70, 20))
        self.line_3.setGeometry(QtCore.QRect(125, 427, 105, 1))
        self.label_18.setGeometry(QtCore.QRect(125, 407, 105, 19))
        self.label_19.setGeometry(QtCore.QRect(390, 420, 105, 19))
        self.listWidget.setGeometry(QtCore.QRect(2, 450, 496, 348))

    def myListWidgetValue(self):
        '''
        listwidget，获取选中的数据
        :return:
        '''
        text_list = self.listWidget.selectedItems()
        return text_list

    def keyPressEvent(self, e):  # 设置一个键盘事件，按键盘上的ESC键，窗体关闭。
        '''
        按下delete后删除选中的数据
        :param e:
        :return:
        '''
        if e.key() == Qt.Key_Delete:
            for i in list(self.myListWidgetValue()):
                # item = self.listWidget.currentItem()
                self.listWidget.takeItem(self.listWidget.row(i))

    '''从brwoseryaml文件读取浏览器信息'''
    def getBrowserType(self):
        browsers = []
        for key, value in self.parseyaml.ReadGuiSelectValue('BrowserType').items():
            browsers.append(value)
        return browsers

    '''编写说明事件'''
    def explainclick(self):
        # 选择保存测试用例的路径
        self.explain_path = QFileDialog.getExistingDirectory(self, "请选择编写说明保存路径")
        explainPath = EXCELTEMPLATE_PATH + '测试用例编写说明.doc'
        # 判断该文件是否已存在该路径下
        if self.explain_path != '':
            if os.path.exists(self.explain_path + '\\测试用例编写说明.doc'):
                shutil.copy2(explainPath, self.explain_path + '\\测试用例编写说明(%s).doc' % self.explainum)
                self.explainum = self.explainum + 1
            else:
                self.explainum = 1
                shutil.copy2(explainPath, self.explain_path + '\\测试用例编写说明.doc')
            QMessageBox.about(self, "提示", "用例编写说明保存成功！")

    '''测试报告事件'''
    def btnstate(self, btn):
        if btn.text() == "不生成":
            if btn.isChecked() == True:
                # 恢复控制初始位置
                self.initialposition()
                # 不生成测试报告，隐藏提示和路径label
                self.label_8.setHidden(True)
                self.label_9.setHidden(True)

        if btn.text() == "生成":
            if btn.isChecked() == True:
                # 调整控制位置
                self.changelocation()
                # 生成测试报告，显示提示和路径label
                self.label_8.setHidden(False)
                self.label_9.setHidden(False)
                # 选择报告保存路径
                self.path = QFileDialog.getExistingDirectory(self, "请选择测试报告保存路径")
                self.path = self.path.replace('/', '\\')
                self.label_9.setText(self.path)
                # 取消或关闭选择窗口
                if self.path == '':
                    # 恢复控制初始位置
                    self.initialposition()
                    # 不生成测试报告，隐藏提示和路径label
                    self.label_8.setHidden(True)
                    self.label_9.setHidden(True)
                    # 选中不生成测试报告
                    self.radioButton_2.setChecked(True)

    '''导入测试用例'''
    def importclick(self):
        try:
            self.CaseNum = []
            # 导入文件名称
            self.filepath, filetype = QFileDialog.getOpenFileName(self, "请导入测试用例", '', 'Excel files(*.xlsx)')
            if self.filepath == '':
                if self.label_10.text() == '请选择excel.xlsx格式表格导入':
                    self.filepath = self.newfile
                    self.label_10.setText('请选择excel.xlsx格式表格导入')
            else:
                self.newfile = self.filepath
                self.newfile = self.filepath.replace('/', '\\')
                # 显示文件名称
                self.label_10.setText(os.path.basename(self.newfile))
                # 打开excel
                self.book = load_workbook(self.newfile)
                # 获取工作表标签
                self.sheets = self.book.sheetnames
                self.CaseSheet = self.sheets[0]
                # 判断是否为测试用例模板
                isCase = ParseExcel(self.filepath).getRowValue(self.CaseSheet, 2)
                isCase1 = ['序号', '用例编号', '用例工作表', '用例标题', '预期结果', '是否执行', '执行结束时间']
                if set(isCase1) < set(isCase):
                    # 获取用例总数，只计算运行的用例
                    isimplement = ParseExcel(self.newfile).getColumnValue(self.CaseSheet, testCase_Isimplement)
                    for i,v in enumerate(isimplement):
                        if v == 'y':
                            self.CaseNum.append(ParseExcel(self.newfile).getCellValue(self.CaseSheet, i+2, testCase_Sheet))
                    # 清空
                    self.comboBox_2.clear()
                    self.comboBox_2.addItem('全部')
                    # 把工作表标签赋值到用例模块选择
                    # self.sheets = self.sheets.pop(0)
                    if len(self.sheets) > 1:
                        del(self.sheets[0])
                    self.comboBox_2.addItems(self.sheets)
                    # 导入用例后先显示全部的用例数量
                    CaseNums = len(self.CaseNum) * int(self.spinBox.text())
                    self.label_15.setText(str(CaseNums))
                else:
                    QMessageBox.about(self, "提示", "用例格式错误，请重新导入用例")
        except Exception as e:
            print(e)

    def ComboxValue(self):
        '''
        下拉框事件，有模板导入时选择模块进行用例显示
        :param tag:
        :return:
        '''
        if self.label_10.text() != '请选择excel.xlsx格式表格导入':
            # 总用例数显示
            if self.comboBox_2.currentText() == '全部':
                # 获取用例总数
                CaseNums = len(self.CaseNum) * int(self.spinBox.text())
                self.label_15.setText(str(CaseNums))
            else:
                self.CaseList = []
                for i in self.CaseNum:
                    if i == self.comboBox_2.currentText():
                        self.CaseList.append(i)
                self.label_15.setText(str(len(self.CaseList)))

    '''导出用例模板'''
    def exportclick(self):
        try:
            # 选择保存测试用例的路径
            self.file_path = QFileDialog.getExistingDirectory(self, "请选择用例模板保存路径")
            excelPath = EXCELTEMPLATE_PATH + '测试用例模板.xlsx'
            # 判断该文件是否已存在该路径下
            if self.file_path != '':
                if os.path.exists(self.file_path + '\\测试用例模板.xlsx'):
                    shutil.copy2(excelPath, self.file_path + '\\测试用例模板(%s).xlsx' % self.filenum)
                    self.filenum = self.filenum+1
                else:
                    self.filenum = 1
                    shutil.copy2(excelPath, self.file_path + '\\测试用例模板.xlsx')
                QMessageBox.about(self, "提示", "用例模板导出成功！")
        except Exception as e:
            print(e)

    def ce(self):
        print(self.CaseNum)
        print(self.comboBox_2.currentText())

    def RunTest(self):
        """
        将gui界面的参数传入yaml文件中,运行脚本
        :return:
        """
        try:
            # logger = Logger('logger').getlog()
            parameter = CONFIG_PATH + 'Parameter.yaml'
            if self.lineEdit.text() == '' or self.lineEdit.text() is None:
                QMessageBox.about(self, "提示", "IP地址不能为空")
            elif not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$", self.lineEdit.text()):
                QMessageBox.about(self, "提示", "请输入正确IP地址")
            elif self.newfile == '':
                QMessageBox.about(self, "提示", "请导入测试用例")
            else:
                try:
                    ParseExcel(self.filepath).wb[self.CaseSheet].cell(50, 50, '测试')
                    ParseExcel(self.filepath).wb.save(self.filepath)
                except PermissionError:
                    QMessageBox.about(self, "提示", "请先关闭用例文件，再运行测试用例")
                    raise
                # # 总用例数显示
                # if self.comboBox_2.currentText() == '全部':
                #     # 获取用例总数
                #     CaseNums = len(self.CaseNum) * int(self.spinBox.text())
                #     self.label_15.setText(str(CaseNums))
                # else:
                #     self.CaseList = []
                #     for i in self.CaseNum:
                #         if i == self.comboBox_2.currentText():
                #             self.CaseList.append(i)
                #     self.label_15.setText(str(len(self.CaseList)))
                if len(self.CaseNum) != 0 or len(self.CaseList) != 0:
                    if self.comboBox.currentText() == 'Google Chrome':
                        # 获取浏览器版本号
                        # self.listWidget.insertItem(0, '获取浏览器版本号中，请稍等...')
                        self.driver = webdriver.Chrome(
                            executable_path=DRIVERS_PATH + 'chrome\\' + '70.0.3538.97\\chromedriver.exe')
                        self.driver.get('chrome://version/')
                        self.version = self.driver.find_element_by_css_selector('#version > span:nth-child(1)').text
                        self.driver.quit()
                        num = ['70', '71', '72', '73', '74', '75', '76', '77', '78']
                        for i in num:
                            if i == self.version[:2]:
                                YamlWrite().Write_Yaml_Updata(CONFIG_PATH + 'Parameter.yaml', 'Version',
                                                              self.version[:2])  # 将value写入ip.yaml文件中
                                # 重置时间
                                self.H = 0
                                self.M = 0
                                self.S = 0
                                # 重置日志显示框
                                self.listWidget.clear()
                                # 清除一些稍后写入的数据
                                YamlWrite().Write_Yaml_Updata(parameter, 'CaseNum', 0)
                                # 写入IP
                                YamlWrite().Write_Yaml_Updata(parameter, 'IP', self.lineEdit.text())
                                # 写入浏览器类型
                                YamlWrite().Write_Yaml_Updata(parameter, 'Browser', self.comboBox.currentText())
                                # 写入是否生成测试报告
                                YamlWrite().Write_Yaml_Updata(parameter, 'ReportAddress', self.label_9.text())
                                # 写入循环次数
                                YamlWrite().Write_Yaml_Updata(parameter, 'loop', self.spinBox.text())
                                # 获取导入的测试用例路径
                                YamlWrite().Write_Yaml_Updata(parameter, 'ImportAddress', self.newfile)
                                # 获取模块信息
                                YamlWrite().Write_Yaml_Updata(parameter, 'Moudle', self.comboBox_2.currentText())
                                # 多线程启动
                                # 已执行用例数线程
                                # self.case = threading.Thread(target=self.display_case)
                                # 运行用例线程
                                if self.radioButton.isChecked() is True:
                                    self.testcase = threading.Thread(target=TestPaperless().RunReport)
                                if self.radioButton_2.isChecked() is True:
                                    self.testcase = threading.Thread(target=TestPaperless().TestCase)
                                # 时间，用例运行计数
                                self.cttime = threading.Thread(target=self.CaseTime)
                                # # 输出显示
                                # self.output = threading.Thread(target=self.OutPut)
                                # if self.Case_num == 0:
                                #     self.case.setDaemon(True)
                                #     self.case.start()
                                if self.testcase.isAlive() is False:
                                    self.testcase.setDaemon(True)
                                    self.testcase.start()
                                # if self.OutPut_num == 0:
                                #     self.output.setDaemon(True)
                                #     self.output.start()
                                self.timer.timeout.connect(self.OutPut)
                                self.timer.start(200)
                                # 启动 按钮点击后，禁用，并置灰
                                self.toolButton_3.setEnabled(False)
                                self.toolButton_3.setStyleSheet(
                                    "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
                                # 暂停 按钮变成蓝色，启用
                                self.toolButton_4.setEnabled(True)
                                self.toolButton_4.setStyleSheet(
                                    "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
                                # 导入 按钮禁用 置灰
                                self.toolButton.setEnabled(False)
                                self.toolButton.setStyleSheet(
                                    "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
                                self.cttime.setDaemon(True)
                                self.cttime.start()
                                break
                            elif int(self.version[:2]) < 70 or int(self.version[:2]) > 78:
                                QMessageBox.about(self, "提示", "浏览器版本不符合，请更新浏览器版本号")
                                break
                    else:
                        # 重置时间
                        self.H = 0
                        self.M = 0
                        self.S = 0
                        # 重置日志显示框
                        self.listWidget.clear()
                        # 清除一些稍后写入的数据
                        YamlWrite().Write_Yaml_Updata(parameter, 'CaseNum', 0)
                        # 写入IP
                        YamlWrite().Write_Yaml_Updata(parameter, 'IP', self.lineEdit.text())
                        # 写入浏览器类型
                        YamlWrite().Write_Yaml_Updata(parameter, 'Browser', self.comboBox.currentText())
                        # 写入是否生成测试报告
                        YamlWrite().Write_Yaml_Updata(parameter, 'ReportAddress', self.label_9.text())
                        # 写入循环次数
                        YamlWrite().Write_Yaml_Updata(parameter, 'loop', self.spinBox.text())
                        # 获取导入的测试用例路径
                        YamlWrite().Write_Yaml_Updata(parameter, 'ImportAddress', self.newfile)
                        # 获取模块信息
                        YamlWrite().Write_Yaml_Updata(parameter, 'Moudle', self.comboBox_2.currentText())
                        # 多线程启动
                        # 已执行用例数线程
                        # self.case = threading.Thread(target=self.display_case)
                        # 运行用例线程
                        if self.radioButton.isChecked() is True:
                            self.testcase = threading.Thread(target=TestPaperless().RunReport)
                        if self.radioButton_2.isChecked() is True:
                            self.testcase = threading.Thread(target=TestPaperless().TestCase)
                        # 时间，用例运行计数
                        self.cttime = threading.Thread(target=self.CaseTime)
                        # # 输出显示
                        # self.output = threading.Thread(target=self.OutPut)
                        # if self.Case_num == 0:
                        #     self.case.setDaemon(True)
                        #     self.case.start()
                        if self.testcase.isAlive() is False:
                            self.testcase.setDaemon(True)
                            self.testcase.start()
                        # if self.OutPut_num == 0:
                        #     self.output.setDaemon(True)
                        #     self.output.start()
                        self.timer.timeout.connect(self.OutPut)
                        self.timer.start(200)
                        # 启动 按钮点击后，禁用，并置灰
                        self.toolButton_3.setEnabled(False)
                        self.toolButton_3.setStyleSheet(
                            "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
                        # 暂停 按钮变成蓝色，启用
                        self.toolButton_4.setEnabled(True)
                        self.toolButton_4.setStyleSheet(
                            "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
                        # 导入 按钮禁用 置灰
                        self.toolButton.setEnabled(False)
                        self.toolButton.setStyleSheet(
                            "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
                        self.cttime.setDaemon(True)
                        self.cttime.start()
                else:
                    QMessageBox.about(self, '提示', '可运行用例数为0，请检查用例')
        except Exception as e:
            print(e)

    def display_case(self):
        '''
        已执行数量
        :return:
        '''
        try:
            if self.testcase.isAlive() is True:
                self.all_already_case = ParseYaml().ReadParameter('CaseNum')
                self.label_17.setText(str(self.all_already_case))
            if self.Case_num == 0:
                self.Case_num += 1
            self.d_times = Timer(1, self.display_case)
            self.d_times.start()
        except Exception as e:
            print('数组越界')

    def Cttime(self):
        '''
        持续时间显示
        :return:
        '''
        try:
            if self.toolButton_3.isEnabled() is False:
                self.S = self.S + 1
                if self.S == 60:
                    self.S = 0
                    self.M = self.M + 1
                    if self.M == 60:
                        self.M = 0
                        self.H = self.H + 1
                if self.M < 10:
                    self.label_18.setText(str(self.H) + ':0' + str(self.M) + ':' + str(self.S))
                else:
                    self.label_18.setText(str(self.H) + ':' + str(self.M) + ':' + str(self.S))
            if self.testcase.isAlive() is False and self.toolButton.isEnabled() is False:
                # 启动 按钮启用
                self.toolButton_3.setEnabled(True)
                self.toolButton_3.setStyleSheet(
                        "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")

                self.toolButton_3.disconnect()
                self.toolButton_3.clicked.connect(self.RunTest)
                # 暂停 按钮禁用
                self.toolButton_4.setEnabled(False)
                self.toolButton_4.setStyleSheet(
                        "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
                # 导入 按钮启用
                self.toolButton.setEnabled(True)
                self.toolButton.setStyleSheet(
                    "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
            if self.Cttime_num == 0:
                self.Cttime_num += 1
            self.c_times = Timer(1, self.Cttime)
            self.c_times.start()
        except Exception as e:
            print(e)

    def Suspend(self):
        """
        暂停用例运行，需要运行完当期用例
        :return:
        """
        parameter = CONFIG_PATH + 'Parameter.yaml'
        YamlWrite().Write_Yaml_Updata(parameter, 'IP', '暂停运行')
        # 启动 按钮点击后，启用
        self.toolButton_3.setEnabled(True)
        self.toolButton_3.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
        self.toolButton_3.disconnect()
        self.toolButton_3.clicked.connect(self.Continue)
        # 暂停 按钮变成灰色，禁用用
        self.toolButton_4.setEnabled(False)
        self.toolButton_4.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")

    def Continue(self):
        """
        继续用例的运行
        :return:
        """
        parameter = CONFIG_PATH + 'Parameter.yaml'
        YamlWrite().Write_Yaml_Updata(parameter, 'IP', self.lineEdit.text())
        # 启动 按钮点击后，禁用，并置灰
        self.toolButton_3.setEnabled(False)
        self.toolButton_3.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
        # 暂停 按钮变成蓝色，启用
        self.toolButton_4.setEnabled(True)
        self.toolButton_4.setStyleSheet(
                "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")


    def OutPut(self):
        """
        持续输出日志
        :return:
        """
        try:
            base_dir = LOGS_PATH
            l = os.listdir(base_dir)
            l.sort(key=lambda fn: os.path.getmtime(base_dir + fn)
            if not os.path.isdir(base_dir + fn) else 0)
            if 'logger' not in l[-1]:
                print('')
            else:
                logpath = os.path.join(base_dir, l[-1])
                filesize = os.path.getsize(logpath)
                blocksize = 1024
                dat_file = open(logpath, 'rb')
                last_line = ""
                if filesize > blocksize:
                    maxseekpoint = (filesize // blocksize)
                    dat_file.seek((maxseekpoint - 1) * blocksize)
                elif filesize:
                    dat_file.seek(0, 0)
                lines = dat_file.readlines()
                if lines:
                    last_line = lines[-1].strip()
                dat_file.close()
                if last_line == '' or last_line is None:
                    print('')
                else:
                    widgetres = []
                    # 获取listwidget中条目数
                    count = self.listWidget.count()
                    # 遍历listwidget中的内容
                    for i in range(count):
                        widgetres.append(self.listWidget.item(i).text())
                    if last_line.decode('utf-8') not in widgetres:
                        self.listWidget.insertItem(len(widgetres), last_line.decode('utf-8'))
                        self.listWidget.scrollToBottom()
                if self.OutPut_num == 0:
                    self.OutPut_num += 1
                if self.testcase.isAlive() is True and self.cttime.isAlive() is False and self.toolButton_3.isEnabled() is False:
                    self.cttime = threading.Thread(target=self.CaseTime)
                    self.cttime.start()
                # self.o_times = Timer(0.2, self.OutPut)
                # self.o_times.start()
        except Exception as e:
            print(e)

    def closeEvent(self, event):
        """
        重写closeEvent方法，实现dialog窗体关闭时执行一些代码
        :param event: close()触发的事件
        :return: None
        """
        if self.toolButton_4.isEnabled() is True:
            QMessageBox.about(self, "提示", "请暂停用例运行后再退出")
            event.ignore()
        else:
            box = QMessageBox(QMessageBox.Question, self.tr("提示"), self.tr("您确定要退出吗？"), QMessageBox.NoButton, self)
            yr_btn = box.addButton(self.tr("确定"), QMessageBox.YesRole)
            box.addButton(self.tr("取消"), QMessageBox.NoRole)
            box.exec_()
            if box.clickedButton() == yr_btn:
                os._exit(5)
                event.accept()
            else:
                event.ignore()

    def CaseTime(self):
        """
        定时器，用于记录用例已运行数量，脚本运行时间，日志输出等功能
        :return:
        """
        try:
            while self.toolButton_3.isEnabled() is False:
                # 用例已运行数量（每隔一秒读取一次用例运行数量）
                # 用例在运行时才进行文件读取
                if self.testcase.isAlive() is True:
                    # 读取用例已运行数量
                    self.all_already_case = ParseYaml().ReadParameter('CaseNum')
                    self.label_17.setText(str(self.all_already_case))
                # 脚本运行时间计数
                # 启动控件在处于禁用状态时开始进行脚本计时（由于有暂停功能，暂停时，时间停止）
                if self.toolButton_3.isEnabled() is False:
                    # 每隔一秒+1
                    self.S = self.S + 1
                    # 当秒计数等于60秒时，分计数+1
                    if self.S == 60:
                        self.S = 0
                        self.M = self.M + 1
                        # 同上，时计数+1
                        if self.M == 60:
                            self.M = 0
                            self.H = self.H + 1
                    if self.M < 10:
                        self.label_18.setText(str(self.H) + ':0' + str(self.M) + ':' + str(self.S))
                    else:
                        self.label_18.setText(str(self.H) + ':' + str(self.M) + ':' + str(self.S))
                # 用于监测用例是否结束运行，结束后改变控件的状态
                if self.testcase.isAlive() is False and self.toolButton_3.isEnabled() is False:
                    # 启动 按钮启用
                    self.toolButton_3.setEnabled(True)
                    self.toolButton_3.setStyleSheet(
                            "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
                    self.toolButton_3.disconnect()
                    self.toolButton_3.clicked.connect(self.RunTest)
                    # 暂停 按钮禁用
                    self.toolButton_4.setEnabled(False)
                    self.toolButton_4.setStyleSheet(
                        "QToolButton{font:12pt '宋体';background-color:#8e7f7c;color: white;border-radius:3px;}")
                    # 导入 按钮启用
                    self.toolButton.setEnabled(True)
                    self.toolButton.setStyleSheet(
                        "QToolButton{font:12pt '宋体';background-color:#419AD9;color: white;border-radius:3px;}")
                time.sleep(1)
        except Exception as e:
            print(e)

if  __name__=="__main__":
    app = 0
    app = QtWidgets.QApplication(sys.argv)
    widget = Application()
    widget.show()
    app.exec_()
    sys.exit(app.exec_())
