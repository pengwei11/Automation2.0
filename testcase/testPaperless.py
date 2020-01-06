#!/usr/bin/env python
# encoding: utf-8
"""
@contact: 1249294960@qq.com
@software: pengwei
@file: testPaperless.py
@time: 2019/11/15 10:38
@desc:
"""

from Utils.ParseExcel import ParseExcel
from Utils.ConfigRead import *
from action.PageAction import PageAction
from selenium.common.exceptions import *   # 导入所有异常类
from Utils.Logger import Logger
from Utils.ParseYaml import ParseYaml
from Utils.WriteFile import YamlWrite
from testcase.HTMLTestRunner_cn import HTMLTestRunner
from datetime import datetime
from openpyxl.styles import Font
import traceback
import time
import re
import unittest
logger = Logger('logger').getlog()


class TestPaperless(unittest.TestCase):

    def setUp(self):
        self.parseyaml = ParseYaml()
        self.testdata_path = self.parseyaml.ReadParameter('ImportAddress')
        self.parseexcel = ParseExcel(self.testdata_path)
        self.pageaction = PageAction()
        self.sheetnames = self.parseexcel.wb.sheetnames
        self.parameter = CONFIG_PATH + 'Parameter.yaml'
        self.CaseNum = 0
        # 创建六个字典，分别储存步骤测试结果，用例测试结果，用例测试时间，错误信息，截图信息，步骤测试时间
        self.time_dic = {}
        self.result_dic = {}
        self.error_dic = {}
        self.picture_dic = {}
        self.caseResult_dic = {}
        self.caseTime_dic = {}
        self.font = Font(color=None)

    def TestCase(self):
        try:
            self.setUp()
            # 获取循环次数
            loop = int(self.parseyaml.ReadParameter('loop'))
            # 获取模块名
            moudle = self.parseyaml.ReadParameter('Moudle')
            # 清除用例旧数据
            self.parseexcel.clearCaseColumnValue(self.sheetnames[0])
            # 清除步骤旧数据
            for i, v in enumerate(self.sheetnames):
                if i == 0:
                    continue
                else:
                    self.parseexcel.clearStepColumnValue(v)
            for l in range(loop):
                # 用例运行数
                try:
                    # 获取'是否执行'列
                    isimplement = self.parseexcel.getColumnValue(self.sheetnames[0], testCase_Isimplement)
                    # 循环'是否执行'列
                    # 如果执行，且模块名符合，则获取用例编号，并切换到对应的工作表，执行用例
                    for index, value in enumerate(isimplement):
                        if moudle == '全部':
                            pd = "value.lower() == 'y'"
                        else:
                            pd = 'value.lower() == "y" and moudle ' \
                                 '== self.parseexcel.getCellValue(self.sheetnames[0], index + 2, testCase_Sheet)'
                        try:
                            # 如果是否执行为空则跳过执行
                            if value is None or value == '':
                                continue
                            elif eval(pd):
                                # 根据'是否执行'；列索引获取对应的工作表名
                                sheetname = self.parseexcel.getCellValue(self.sheetnames[0], index + 2, testCase_Sheet)
                                # 根据'是否执行'列索引获取对应的用例编号
                                testcasenum = self.parseexcel.getCellValue(self.sheetnames[0], index + 2, testCase_Num)
                                # 切换到用例对应的工作表
                                # sheetnames = self.parseexcel.wb[sheetname]
                                '''
                                根据用例编号(testcasenum)获取预置条件编号
                                '''
                                # 获取用例步骤的用例编号类，并执行对应用例编号的用例步骤（增加表内是否有合并单元格的判断）
                                if self.parseexcel.ismerge(sheetname):
                                    teststepnum = self.parseexcel.getMergeColumnValue(sheetname, testStep_Num)
                                else:
                                    teststepnum = self.parseexcel.getColumnValue(sheetname, testStep_Num)
                                # 循环用例步骤编号，根据索引获取预置条件编号
                                testPrenum = ''
                                for i, v in enumerate(teststepnum):
                                    if v == testcasenum:
                                        # 用例前置条件编号
                                        testPrenum = self.parseexcel.getCellValue(sheetname, i + 2, testStep_Preset)
                                        break

                                # 循环用例步骤编号，找到与预置条件编号相同的用例步骤编号
                                # 循环所有的步骤编号
                                # 获取对应用例编号的步骤编号的关键字，定位方式，表达式，操作值
                                url = ParseYaml().ReadParameter('IP')
                                while re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$", url) is None and re.match(r'[^\s]*[.com|.cn]', url) is None:
                                    url = ParseYaml().ReadParameter('IP')  # 从输入框获取浏览器地址
                                    # 先打开浏览器，进入指定IP地址
                                    time.sleep(1)
                                self.pageaction.openBrowser()
                                self.pageaction.getUrl('http://%s' % url)
                                # 执行预置条件
                                for t, v in enumerate(teststepnum):
                                    if v == testPrenum:
                                        # 用例执行步骤
                                        pre_stepname = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Describe)
                                        # 获取预置条件关键字
                                        pre_keyword = self.parseexcel.getCellValue(sheetname, t + 2, testStep_KeyWord)
                                        # 去除前后空格
                                        if pre_keyword is not None:
                                            pre_keyword = pre_keyword.strip()
                                        # 获取定位方式
                                        pre_location = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Location)
                                        # 去除前后空格
                                        if pre_location is not None:
                                            pre_location = pre_location.strip()
                                        # 获取定位表达式
                                        pre_locator = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Locator)
                                        if type(pre_locator) is int:
                                            pre_locator = str(self.parseexcel.getCellValue(sheetname, t + 2, testStep_Locator))
                                        # 获取输入值
                                        pre_testvalue = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Value)
                                        # 如果输入值为 int 类型，则强转为 str 类型，用于字符串拼接
                                        if pre_testvalue is not None and type(pre_testvalue) is not str:
                                            pre_testvalue = str(self.parseexcel.getCellValue(sheetname, t + 2, testStep_Value))
                                        # 总共有四种情况可以正常执行，其他情况则会将用例判断为运行失败
                                        # 1.关键字，定位方式，表达式，输入值全部不为空的情况 例：send_keyslower
                                        # 2.关键字，输入值不为空，定位方式，表达式为空的情况 例：assert（断言）
                                        # 3.关键字，定位方式，表达式不为空，输入值为空的情况 例：click
                                        # 4.关键字不为空，定位方式，表达式，输入值为空的情况 例 getTitle
                                        if pre_keyword and pre_location and pre_locator and pre_testvalue:
                                            pre_fun = 'self.pageaction' + '.' + pre_keyword + '(' + '"' + pre_location + '"' + ', ' + '"' + pre_locator + '"' + ', ' + '"' + \
                                                      pre_testvalue + '"' + ')'
                                        elif pre_keyword and pre_testvalue and pre_location is None or pre_location == '' \
                                                and pre_locator is None or pre_location == '':
                                            pre_fun = 'self.pageaction' + '.' + pre_keyword + '(' + '"' + pre_testvalue + '"' + ')'
                                        elif pre_keyword and pre_location and pre_locator and pre_testvalue is None or pre_testvalue == '':
                                            pre_fun = 'self.pageaction' + '.' + pre_keyword + '(' + '"' + pre_location + '"' + ', ' + '"' + pre_locator + '"' + ')'
                                        elif pre_keyword and pre_location is None or pre_location == '' and pre_locator is None \
                                                or pre_locator == '' and pre_testvalue is None or pre_testvalue == '':
                                            pre_fun = 'self.pageaction' + '.' + pre_keyword + '(' + ')'
                                        elif pre_keyword is None or pre_keyword == '' and pre_location is None or pre_location == '' \
                                                and pre_locator is None or pre_locator == '' and pre_testvalue is None or pre_testvalue == '':
                                            continue
                                        else:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t+2] = '关键字对应参数错误'
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '关键字对应参数错误')
                                            logger.info('关键字对应参数错误')
                                            print('关键字对应参数错误')
                                            continue
                                        # 执行用例
                                        try:
                                            # eval 将字符串转换为可执行的python语句
                                            eval(pre_fun)
                                        # 抛出异常的情况，将失败结果写入excel表格中
                                        except TypeError:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t+2] = '关键字参数个数错误，请检查参数'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '关键字参数个数错误，请检查参数')
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            logger.info('关键字参数个数错误，请检查参数')
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                            print('关键字参数个数错误，请检查参数')
                                        except TimeoutException:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t+2] = '元素定位超时，' \
                                                                                            '请检查上一步是否执行成功，或元素定位方式'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '元素定位超时，'
                                            #                                                             '请检查上一步是否执行成功，或元素定位方式')
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            logger.info('元素定位超时，请检查上一步是否执行成功，或元素定位方式')
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                            print('元素定位超时，请检查上一步是否执行成功，或元素定位方式')
                                        except TimeoutError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '断言失败'
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                            logger.info(e)
                                        except AttributeError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '元素定位超时，请检查元素定位'
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                            logger.info(e)
                                        except AssertionError:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Failed'
                                            self.error_dic.setdefault(sheetname, {})[t+2] = '断言失败'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Failed')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '断言失败')
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                        except WebDriverException:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t+2] = '浏览器异常，' \
                                                                                            '请检查浏览器驱动或运行过程中是否被强制关闭'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '浏览器异常，'
                                            #                                                             '请检查浏览器驱动或运行过程中是否被强制关闭')
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            logger.info('浏览器异常，请检查浏览器驱动或运行过程中是否被强制关闭')
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                            print('浏览器异常，请检查浏览器驱动或运行过程中是否被强制关闭')
                                        except Exception:
                                            error_info = traceback.format_exc()
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t+2] = error_info
                                            # # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, error_info)
                                            logger.info('步骤"{}"执行失败'.format(pre_stepname))
                                            print('步骤"{}"执行失败'.format(pre_stepname))
                                        else:
                                            # 写入测试结果
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t+2] = 'Pass'
                                            # self.parseexcel.writeCellValue(sheetname, t + 2, testStep_Result+l, 'Pass')
                                            logger.info('步骤"{}"执行成功'.format(pre_stepname))
                                            print('步骤"{}"执行成功'.format(pre_stepname))
                                        finally:
                                            # 截图
                                            pic = self.pageaction.saveScreeShot(sheetname, testcasenum)
                                            # 将截图信息以及测试时间存入字典中
                                            Time = datetime.now()
                                            Time.strftime('%Y:%m:%d %H:%M:%S')
                                            self.picture_dic.setdefault(sheetname, {})[t+2] = pic
                                            self.time_dic.setdefault(sheetname, {})[t + 2] = Time
                                            # self.parseexcel.writeCellValue(sheetname, t + 2, testStep_Picture, pic)
                                            # self.parseexcel.writeCellTime(sheetname, t + 2, testStep_EndTime)
                                    else:
                                        continue

                                # 将用例步骤工作表内的用例编号以字典的方式循环
                                for t, v in enumerate(teststepnum):
                                    # 用例步骤（用例编号） 与 用例列表（用例编号）相同的
                                    if v == testcasenum:
                                        # 用例执行步骤
                                        stepname = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Describe)
                                        # 获取关键字
                                        keyword = self.parseexcel.getCellValue(sheetname, t + 2, testStep_KeyWord)
                                        # 去除前后空格
                                        if keyword is not None:
                                            keyword = keyword.strip()
                                        # 获取定位方式
                                        location = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Location)
                                        # 去除前后空格
                                        if location is not None:
                                            location = location.strip()
                                        # 获取定位表达式
                                        locator = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Locator)
                                        if type(locator) is int:
                                            locator = str(self.parseexcel.getCellValue(sheetname, t + 2, testStep_Locator))
                                        # 获取输入值
                                        testvalue = self.parseexcel.getCellValue(sheetname, t + 2, testStep_Value)
                                        # 如果输入值为 int 类型，则强转为 str 类型，用于字符串拼接
                                        if testvalue is not None and type(testvalue) is not str:
                                            testvalue = str(self.parseexcel.getCellValue(sheetname, t + 2, testStep_Value))
                                        # if testvalue.lower() == 'none':
                                        #     testvalue = ''
                                        # 进行关键字拼接
                                        # 总共有四种情况可以正常执行，其他情况则会将用例判断为运行失败
                                        # 1.关键字，定位方式，表达式，输入值全部不为空的情况 例：send_keys
                                        # 2.关键字，输入值不为空，定位方式，表达式为空的情况 例：assert（断言）
                                        # 3.关键字，定位方式，表达式不为空，输入值为空的情况 例：click
                                        # 4.关键字不为空，定位方式，表达式，输入值为空的情况 例 getTitle
                                        if keyword and location and locator and testvalue:
                                            fun = 'self.pageaction' + '.' + keyword + '(' + '"' + location + '"' + ', ' + '"' + locator + '"' + ', ' + '"' + \
                                                  testvalue + '"' + ')'
                                        elif keyword and testvalue and location is None or location == '' \
                                                and locator is None or location == '':
                                            fun = 'self.pageaction' + '.' + keyword + '(' + '"' + testvalue + '"' + ')'
                                        elif keyword and location and locator and testvalue is None or testvalue == '':
                                            fun = 'self.pageaction' + '.' + keyword + '(' + '"' + location + '"' + ', ' + '"' + locator + '"' + ')'
                                        elif keyword and location is None or location == '' and locator is None \
                                                or locator == '' and testvalue is None or testvalue == '':
                                            fun = 'self.pageaction' + '.' + keyword + '(' + ')'
                                        elif keyword is None or keyword == '' and location is None or location == '' \
                                                and locator is None or locator == '' and testvalue is None or testvalue == '':
                                            continue
                                        else:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '关键字对应参数错误'
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '关键字对应参数错误')
                                            logger.info('关键字对应参数错误')
                                            print('关键字对应参数错误')
                                            continue
                                        # 执行用例
                                        try:
                                            # eval 将字符串转换为可执行的python语句
                                            eval(fun)
                                        # 抛出异常的情况，将失败结果写入excel表格中
                                        except TypeError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '关键字参数个数错误，请检查参数'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '关键字参数个数错误，请检查参数')
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            logger.info('关键字参数个数错误，请检查参数')
                                            print('步骤"{}"执行失败'.format(stepname))
                                            print('关键字参数个数错误，请检查参数')
                                            logger.info(e)
                                        except TimeoutException as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '元素定位超时，' \
                                                                                              '请检查上一步是否执行成功，或元素定位方式'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '元素定位超时，'
                                            #                                                             '请检查上一步是否执行成功，或元素定位方式')
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            logger.info('元素定位超时，请检查上一步是否执行成功，或元素定位方式')
                                            print('步骤"{}"执行失败'.format(stepname))
                                            print('元素定位超时，请检查上一步是否执行成功，或元素定位方式')
                                            logger.info(e)
                                        except TimeoutError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '断言失败'
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            print('步骤"{}"执行失败'.format(stepname))
                                            logger.info(e)
                                        except AttributeError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '元素定位超时，请检查元素定位'
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            print('步骤"{}"执行失败'.format(stepname))
                                            logger.info(e)
                                        except AssertionError as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Failed'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '断言失败'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Failed')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '断言失败')
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            print('步骤"{}"执行失败'.format(stepname))
                                            logger.info(e)
                                        except WebDriverException as e:
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = '浏览器异常，' \
                                                                                              '请检查浏览器驱动或运行过程中是否被强制关闭'
                                            # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, '浏览器异常，'
                                            #                                                             '请检查浏览器驱动或运行过程中是否被强制关闭')
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            logger.info('浏览器异常，请检查浏览器驱动或运行过程中是否被强制关闭')
                                            print('步骤"{}"执行失败'.format(stepname))
                                            print('浏览器异常，请检查浏览器驱动或运行过程中是否被强制关闭')
                                            logger.info(e)
                                        except Exception:
                                            error_info = traceback.format_exc()
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Skip'
                                            self.error_dic.setdefault(sheetname, {})[t + 2] = error_info
                                            # # 写入测试时间，测试结果，错误信息，错误截图
                                            # self.parseexcel.writeCellValue(sheetname, t+2, testStep_Result+l, 'Skip')
                                            # self.parseexcel.writeCellValues(sheetname, t + 2, error_info)
                                            logger.info('步骤"{}"执行失败'.format(stepname))
                                            print('步骤"{}"执行失败'.format(stepname))
                                        else:
                                            # 写入测试结果
                                            # 将结果以及错误信息存入字典
                                            self.result_dic.setdefault(sheetname, {})[t + 2] = 'Pass'
                                            # self.parseexcel.writeCellValue(sheetname, t + 2, testStep_Result+l, 'Pass')
                                            logger.info('步骤"{}"执行成功'.format(stepname))
                                            print('步骤"{}"执行成功'.format(stepname))
                                        finally:
                                            # 截图
                                            pic = self.pageaction.saveScreeShot(sheetname, testcasenum)
                                            # 将截图信息以及测试时间存入字典中
                                            Time = datetime.now()
                                            Time.strftime('%Y:%m:%d %H:%M:%S')
                                            self.picture_dic.setdefault(sheetname, {})[t + 2] = pic
                                            self.time_dic.setdefault(sheetname, {})[t + 2] = Time
                                            # self.parseexcel.writeCellValue(sheetname, t + 2, testStep_Picture, pic)
                                            # self.parseexcel.writeCellTime(sheetname, t + 2, testStep_EndTime)
                                    else:
                                        continue
                                self.CaseNum += 1
                                YamlWrite().Write_Yaml_Updata(self.parameter, 'CaseNum', self.CaseNum)
                                self.pageaction.quitBrowser()
                                # 写入测试结果
                                for r, v in self.result_dic.items():
                                    for a, b in v.items():
                                        if b == 'Pass':
                                            self.parseexcel.wb[r].cell(int(a), testStep_Result + l).font = Font(color='33ff33')
                                            self.parseexcel.wb[r].cell(int(a), testStep_Result + l, b)
                                        elif b == 'Failed':
                                            self.parseexcel.wb[r].cell(int(a), testStep_Result + l).font = Font(color='cc0000')
                                            self.parseexcel.wb[r].cell(int(a), testStep_Result + l, b)
                                        elif b == 'Skip':
                                            self.parseexcel.wb[r].cell(int(a), testStep_Result + l).font = Font(color='D1D1D1')
                                            self.parseexcel.wb[r].cell(int(a), testStep_Result + l, b)
                                        else:
                                            continue
                                # 通过循环对应 用例编号的步骤的结果，全部为pass的则写入用例Pass，有一条失败的则写入Failed
                                for s, b in enumerate(teststepnum):
                                    # 获取测试结果
                                    if b == testcasenum:
                                        if self.parseexcel.getCellValue(sheetname, s+2, testStep_Result) is None\
                                                or self.parseexcel.getCellValue(sheetname, s+2, testStep_Result) == '':
                                            continue
                                        elif self.parseexcel.getCellValue(sheetname, s+2, testStep_Result) == 'Pass':
                                            # 将用例测试结果存入字典
                                            self.caseResult_dic.setdefault(self.sheetnames[0], {})[index+2] = 'Pass'
                                            # self.parseexcel.writeCellValue(self.sheetnames[0], index+2, testCase_Result+l, 'Pass')
                                        elif self.parseexcel.getCellValue(sheetname, s+2, testStep_Result) == 'Skip':
                                            self.caseResult_dic.setdefault(self.sheetnames[0], {})[index+2] = 'Failed'
                                            break
                                            # self.parseexcel.writeCellValue(self.sheetnames[0], index + 2, testCase_Result+l,
                                            #                                'Failed')
                                        else:
                                            self.caseResult_dic.setdefault(self.sheetnames[0], {})[index+2] = 'Failed'
                                            break
                                            # self.parseexcel.writeCellValue(self.sheetnames[0], index + 2, testCase_Result+l,
                                            #                                'Failed')
                                Time = datetime.now()
                                Time.strftime('%Y:%m:%d %H:%M:%S')
                                # 增加时间写入，以及已运行数量统计
                                self.caseTime_dic.setdefault(self.sheetnames[0], {})[index+2] = Time
                                # self.parseexcel.writeCellTime(self.sheetnames[0], index + 2, testCase_EndTime)
                                # 增加用例时间运行间隔，默认1秒（通过配置文件进行修改）
                                time.sleep(self.parseyaml.ReadTimeWait('casetime'))
                            else:
                                continue
                        except Exception as e:
                            Time = datetime.now()
                            Time.strftime('%Y:%m:%d %H:%M:%S')
                            self.caseResult_dic.setdefault(self.sheetnames[0], {})[index + 2] = 'Failed'
                            self.caseTime_dic.setdefault(self.sheetnames[0], {})[index + 2] = Time
                            # self.parseexcel.writeCellTime(self.sheetnames[0], index + 2, testCase_EndTime)
                            # self.parseexcel.writeCellValue(self.sheetnames[0], index + 2, testCase_Result+l,
                            #                                'Failed')
                            self.CaseNum += 1
                            YamlWrite().Write_Yaml_Updata(self.parameter, 'CaseNum', self.CaseNum)
                            logger.info(e)
                    logger.info('正在写入测试结果，请勿关闭界面...')
                    # 读取所有字典，将结果写入excel中
                    for t, v in self.time_dic.items():
                        for a, b in v.items():
                            self.parseexcel.wb[t].cell(int(a), testStep_EndTime, b)
                    for e, v in self.error_dic.items():
                        for a, b in v.items():
                            self.parseexcel.wb[e].cell(int(a), testStep_Error, b)
                    for p, v in self.picture_dic.items():
                        for a, b in v.items():
                            self.parseexcel.wb[p].cell(int(a), testStep_Picture).value = '=HYPERLINK("{}", "{}")'.format(b, b)
                    for ct, v in self.caseTime_dic.items():
                        for a, b in v.items():
                            self.parseexcel.wb[ct].cell(int(a), testCase_EndTime, b)
                    for cr, v in self.caseResult_dic.items():
                        for a, b in v.items():
                            if b == 'Pass':
                                self.parseexcel.wb[cr].cell(int(a), testCase_Result + l).font = Font(color='33ff33')
                                self.parseexcel.wb[cr].cell(int(a), testCase_Result + l, b)
                            elif b == 'Failed':
                                self.parseexcel.wb[cr].cell(int(a), testCase_Result + l).font = Font(color='cc0000')
                                self.parseexcel.wb[cr].cell(int(a), testCase_Result + l, b)
                            else:
                                continue
                    # 获取excel中'用例工作表'列的不为None的总行数
                    total_case = list(filter(None, self.parseexcel.getColumnValue(self.sheetnames[0], testCase_Sheet)))
                    # 写入excel表的总用例数单元格中
                    self.parseexcel.writeCellValue(self.sheetnames[0], 1, 3, len(total_case)-1)
                    # 循环执行结果列中为pass的列
                    pass_case = []
                    faild_case = []
                    for pi in list(filter(None, self.parseexcel.getColumnValue(self.sheetnames[0], testCase_Result))):
                        if pi.lower() == 'pass':
                            pass_case.append(pi)
                        elif pi.lower() == 'failed':
                            faild_case.append(pi)
                        else:
                            continue
                    # 写入excel表中的通过用力数单元格中
                    self.parseexcel.writeCellValue(self.sheetnames[0], 1, 5, len(pass_case))
                    # 写入excel表中的失败用例数单元格中
                    self.parseexcel.writeCellValue(self.sheetnames[0], 1, 7, len(faild_case))
                    # 循环是否执行列的中n的数量
                    n_case = []
                    for ni in list(filter(None, self.parseexcel.getColumnValue(self.sheetnames[0], testCase_Isimplement))):
                        if ni.lower() == 'n':
                            n_case.append(ni)
                    # 写入excel表中的未测试用例的单元格中
                    self.parseexcel.writeCellValue(self.sheetnames[0], 1, 9, len(n_case))
                except Exception as e:
                    logger.info(e)
                    self.pageaction.quitBrowser()
                finally:
                    self.parseexcel.wb.save(self.testdata_path)
                    logger.info('用例测试结束')
                    print('用例测试结束')
        except Exception as e:
            print(e)
            # 异常结束时，关闭文件流
            self.parseexcel.wb.close()

    def RunReport(self):
        report_path = ParseYaml().ReadParameter('ReportAddress') # 报告存放位置
        timestr = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        filename = report_path+'\\'+timestr+'.html'
        fp = open(filename, 'wb')
        # suites = unittest.defaultTestLoader.discover(TESTCASE_PATH, pattern='test*.py', top_level_dir=TESTCASE_PATH)
        suites = unittest.TestSuite()
        suites.addTest(TestPaperless('TestCase'))
        runner = HTMLTestRunner(
            title='无纸化测试报告',
            description='',
            stream=fp,
            verbosity=2,
        )
        runner.run(suites)
        fp.close()

if __name__ == '__main__':
    TestPaperless().TestCase()